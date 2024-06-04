import datetime

import openpyxl as xl
from docxtpl import DocxTemplate

workbook = xl.load_workbook('test.xlsx')
sheet_1 = workbook['Sheet1']

template = DocxTemplate('template.docx')

table_contents = []
for i in range(2, sheet_1.max_row + 1):
    table_contents.append({
        'Index': i - 1,
        'id': sheet_1.cell(i, 1).value,
        'name': sheet_1.cell(i, 2).value,
        'age': sheet_1.cell(i, 3).value,
        'dep': sheet_1.cell(i, 4).value
    })

context = {
    'title': 'Automated Report',
    'day': datetime.datetime.now().strftime('%d'),
    'month': datetime.datetime.now().strftime('%b'),
    'year': datetime.datetime.now().strftime('%Y'),
    'table_contents': table_contents
}

template.render(context)
template.save('Automated_report.docx')
